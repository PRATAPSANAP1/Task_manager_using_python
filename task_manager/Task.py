import tkinter as tk
from tkinter import messagebox, ttk
from tkcalendar import DateEntry
import openpyxl
import os
from datetime import datetime, timedelta, time

class Task:
    def __init__(self, task_name, description, person_name, priority, due_date, due_time, completed=False):
        self.task_name = task_name
        self.description = description
        self.person_name = person_name
        self.priority = priority
        self.due_date = due_date
        self.due_time = due_time
        self.completed = completed

class TaskManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Task Manager")

        self.tasks = []
        self.completed_tasks = []

        self.task_name_var = tk.StringVar()
        self.description_var = tk.StringVar()
        self.person_name_var = tk.StringVar()
        self.priority_var = tk.StringVar()
        self.due_date_var = tk.StringVar()
        self.due_time_var = tk.StringVar()

        self.create_widgets()
        self.check_due_dates_periodically()

    def create_widgets(self):
        # Create a Notebook widget
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True)

        # Create the task management tab
        task_tab = ttk.Frame(self.notebook)
        self.notebook.add(task_tab, text='Task Management')

        # Task Name Label and Entry
        tk.Label(task_tab, text="Task Name:").grid(row=0, column=0, sticky="w")
        task_name_entry = tk.Entry(task_tab, textvariable=self.task_name_var)
        task_name_entry.grid(row=0, column=1, padx=10, pady=5)

        # Description Label and Entry
        tk.Label(task_tab, text="Description:").grid(row=1, column=0, sticky="w")
        description_entry = tk.Entry(task_tab, textvariable=self.description_var)
        description_entry.grid(row=1, column=1, padx=10, pady=5)

        # Person Name Label and Entry
        tk.Label(task_tab, text="Person Name:").grid(row=2, column=0, sticky="w")
        person_name_entry = tk.Entry(task_tab, textvariable=self.person_name_var)
        person_name_entry.grid(row=2, column=1, padx=10, pady=5)

        # Priority Label and Dropdown
        tk.Label(task_tab, text="Priority:").grid(row=3, column=0, sticky="w")
        priority_values = ["Low", "Medium", "High"]
        priority_dropdown = ttk.Combobox(task_tab, textvariable=self.priority_var, values=priority_values)
        priority_dropdown.grid(row=3, column=1, padx=10, pady=5)

        # Due Date Label and Calendar
        tk.Label(task_tab, text="Due Date:").grid(row=4, column=0, sticky="w")
        due_date_entry = DateEntry(task_tab, textvariable=self.due_date_var, date_pattern="yyyy-mm-dd")
        due_date_entry.grid(row=4, column=1, padx=10, pady=5)

        # Due Time Label and Entry
        tk.Label(task_tab, text="Due Time (HH:MM):").grid(row=5, column=0, sticky="w")
        due_time_entry = tk.Entry(task_tab, textvariable=self.due_time_var)
        due_time_entry.grid(row=5, column=1, padx=10, pady=5)
        due_time_entry.insert(tk.END, "")

        # Add Task Button
        add_task_button = tk.Button(task_tab, text="Add Task", command=self.add_task)
        add_task_button.grid(row=6, column=0, columnspan=2, padx=10, pady=5)

        # Mark as Completed Button
        mark_completed_button = tk.Button(task_tab, text="Mark as Completed", command=self.mark_as_completed)
        mark_completed_button.grid(row=7, column=0, columnspan=2, padx=10, pady=5)

        # Task List Treeview
        self.task_list_treeview = ttk.Treeview(task_tab, columns=("Task Name", "Description", "Person Name", "Priority", "Due Date", "Due Time"))
        self.task_list_treeview.grid(row=8, column=0, columnspan=2, padx=10, pady=5)
        self.task_list_treeview.heading("#0", text="Task Name")
        self.task_list_treeview.heading("#1", text="Description")
        self.task_list_treeview.heading("#2", text="Person Name")
        self.task_list_treeview.heading("#3", text="Priority")
        self.task_list_treeview.heading("#4", text="Due Date")
        self.task_list_treeview.heading("#5", text="Due Time")

        # Create the completed tasks tab
        completed_tab = ttk.Frame(self.notebook)
        self.notebook.add(completed_tab, text='Completed Tasks')

        # Completed Task List Treeview
        self.completed_task_list_treeview = ttk.Treeview(completed_tab, columns=("Task Name", "Description", "Person Name", "Priority", "Due Date", "Due Time"))
        self.completed_task_list_treeview.grid(row=0, column=0, padx=10, pady=5)
        self.completed_task_list_treeview.heading("#0", text="Task Name")
        self.completed_task_list_treeview.heading("#1", text="Description")
        self.completed_task_list_treeview.heading("#2", text="Person Name")
        self.completed_task_list_treeview.heading("#3", text="Priority")
        self.completed_task_list_treeview.heading("#4", text="Due Date")
        self.completed_task_list_treeview.heading("#5", text="Due Time")

        # Delete Task Button
        delete_task_button = tk.Button(task_tab, text="Delete Task", command=self.delete_task)
        delete_task_button.grid(row=9, column=0, padx=10, pady=5, sticky="w")

        # Clear Task Button
        clear_task_button = tk.Button(task_tab, text="Clear Task", command=self.clear_task)
        clear_task_button.grid(row=9, column=1, padx=10, pady=5, sticky="e")

        # Load Tasks Button
        load_button = tk.Button(task_tab, text="Load Tasks", command=self.load_from_excel)
        load_button.grid(row=9, column=0, columnspan=2, padx=10, pady=5)

    def add_task(self):
        task_name = self.task_name_var.get()
        description = self.description_var.get()
        person_name = self.person_name_var.get()
        priority = self.priority_var.get()
        due_date = self.due_date_var.get()
        due_time = self.due_time_var.get()

        if task_name and description and person_name and priority and due_date and due_time:
            task = Task(task_name, description, person_name, priority, due_date, due_time)
            self.tasks.append(task)

            self.task_list_treeview.insert("", tk.END, text=task.task_name, values=(task.description, task.person_name,
                                                                               task.priority, task.due_date, task.due_time))

            self.task_name_var.set("")
            self.description_var.set("")
            self.person_name_var.set("")
            self.priority_var.set("")
            self.due_date_var.set("")
            self.due_time_var.set("")

            # Automatically save the task to the Excel sheet
            self.save_to_excel()
        else:
            messagebox.showerror("Error", "Please fill in all fields.")

    def delete_task(self):
        selected_item = self.task_list_treeview.selection()
        if selected_item:
            task_name = self.task_list_treeview.item(selected_item)["text"]  # Get task name from text
            for task in self.tasks:
                if task.task_name == task_name:
                    self.tasks.remove(task)
                    self.task_list_treeview.delete(selected_item)
                    # Delete task from Excel sheet
                    self.delete_from_excel(task)
                    break

    def delete_from_excel(self, task):
        current_directory = os.getcwd()
        file_path = os.path.join(current_directory, "taskmanager.xlsx")
        if not os.path.exists(file_path):
            messagebox.showerror("Error", "No tasks found. Excel file not found.")
            return

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == task.task_name:  # Check if task name matches
                ws.delete_rows(row, amount=1)  # Delete the row from Excel sheet
                break

        wb.save(file_path)
        messagebox.showinfo("Success", "Task deleted from Excel.")

    def mark_as_completed(self):
        selected_item = self.task_list_treeview.selection()
        if selected_item:
            task_name = self.task_list_treeview.item(selected_item)["text"]  # Get task name from text
            for task in self.tasks:
                if task.task_name == task_name:
                    task.completed = True
                    self.completed_tasks.append(task)
                    self.tasks.remove(task)
                    self.task_list_treeview.delete(selected_item)
                    # Update task status in Excel sheet
                    self.update_status_in_excel(task)
                    self.completed_task_list_treeview.insert("", tk.END, text=task.task_name, values=(task.description, task.person_name,
                                                                                                        task.priority, task.due_date, task.due_time))
                    break

    def update_status_in_excel(self, task):
        current_directory = os.getcwd()
        file_path = os.path.join(current_directory, "taskmanager.xlsx")
        if not os.path.exists(file_path):
            messagebox.showerror("Error", "No tasks found. Excel file not found.")
            return

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == task.task_name:  # Check if task name matches
                ws.cell(row=row, column=7, value="Completed")  # Update the status to "Completed"
                break

        wb.save(file_path)
        messagebox.showinfo("Success", "Task status updated in Excel.")

    def clear_task(self):
        self.task_name_var.set("")
        self.description_var.set("")
        self.person_name_var.set("")
        self.priority_var.set("")
        self.due_date_var.set("")
        self.due_time_var.set("")

    def load_from_excel(self):
        current_directory = os.getcwd()
        file_path = os.path.join(current_directory, "taskmanager.xlsx")
        if not os.path.exists(file_path):
            messagebox.showerror("Error", "No tasks found. Excel file not found.")
            return

        self.clear_task()

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            task_name, description, person_name, priority, due_date, due_time, status = row
            task = Task(task_name, description, person_name, priority, due_date, due_time, completed=status=="Completed")
            if task.completed:
                self.completed_tasks.append(task)
                self.completed_task_list_treeview.insert("", tk.END, text=task.task_name, values=(task.description, task.person_name,
                                                                                                     task.priority, task.due_date, task.due_time))
            else:
                self.tasks.append(task)
                self.task_list_treeview.insert("", tk.END, text=task.task_name, values=(task.description, task.person_name,
                                                                               task.priority, task.due_date, task.due_time))
        messagebox.showinfo("Success", "Tasks loaded from Excel.")

    def save_to_excel(self):
        current_directory = os.getcwd()
        file_path = os.path.join(current_directory, "taskmanager.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Task Name", "Description", "Person Name", "Priority", "Due Date", "Due Time", "Status"])
        for task in self.tasks:
            ws.append([task.task_name, task.description, task.person_name, task.priority, task.due_date, task.due_time, "Pending"])
        for task in self.completed_tasks:
            ws.append([task.task_name, task.description, task.person_name, task.priority, task.due_date, task.due_time, "Completed"])
        wb.save(file_path)
        messagebox.showinfo("Success", f"Data saved to {file_path}.")

    def check_due_dates_periodically(self):
        # Get the current time
        current_time = datetime.now().time()

        # Define the fixed reminder time
        reminder_time = time(hour=8, minute=0)  # Example: Remind at 8:00 AM

        # If the current time is after the reminder time, schedule the reminder for the next day
        if current_time >= reminder_time:
            next_reminder_datetime = datetime.combine(datetime.now().date() + timedelta(days=1), reminder_time)
        else:
            next_reminder_datetime = datetime.combine(datetime.now().date(), reminder_time)

        # Calculate the time until the next reminder
        time_until_next_reminder_seconds = int((next_reminder_datetime - datetime.now()).total_seconds())

        # Schedule the check_due_dates method to be called again at the next reminder time
        self.root.after(time_until_next_reminder_seconds * 1000, self.check_due_dates)

    def check_due_dates(self):
        for task in self.tasks:
            due_date_time = datetime.strptime(f"{task.due_date} {task.due_time}", "%Y-%m-%d %H:%M")
            current_date_time = datetime.now()
            time_until_due = due_date_time - current_date_time

            if time_until_due <= timedelta(days=3):
                messagebox.showwarning("Task Deadline", f"The task '{task.task_name}' is due in {time_until_due}.")
            elif time_until_due <= timedelta(minutes=30):
                messagebox.showinfo("Task Deadline", f"The task '{task.task_name}' is due in {time_until_due}.")
            elif time_until_due <= timedelta(minutes=3):
                messagebox.showinfo("Task Deadline", f"The task '{task.task_name}' is due in {time_until_due}.")

root = tk.Tk()
app = TaskManagerApp(root)
root.mainloop()